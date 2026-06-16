from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


TASK_DIR = Path("<research-root>/example-rag-eval-tooling-xlsx-20260616")
SOURCES_PATH = TASK_DIR / "sources.jsonl"
FINDINGS_PATH = TASK_DIR / "findings.jsonl"
OUTPUT_DIR = TASK_DIR / "workspace" / "outputs" / "rag-eval-tooling-matrix"
WORKBOOK_PATH = OUTPUT_DIR / "rag-eval-tooling-matrix.xlsx"
VALIDATION_PATH = OUTPUT_DIR / "validation-report.md"
GENERATION_SUMMARY_PATH = TASK_DIR / "workspace" / "analysis" / "package-generation-summary.json"

GENERATED_ON = date(2026, 6, 16).isoformat()


TOOLS = [
    {
        "tool": "LangSmith",
        "category": "Managed evaluation and observability platform",
        "recommended_use": "Best default shortlist item for LangChain-heavy teams that need datasets, experiment runs, traces, enterprise deployment options, and governance controls.",
        "evaluation_posture": "Strong managed workflow",
        "deployment_posture": "Cloud, hybrid, and enterprise self-hosted options",
        "rag": 4,
        "ci": 4,
        "dataset": 5,
        "tracing": 5,
        "integration": 5,
        "data_control": 4,
        "governance": 5,
        "evidence": 5,
        "caveat": "Strongest fit when the application stack already uses LangChain or the team accepts a managed/commercial control plane.",
        "source_urls": [
            "https://docs.langchain.com/langsmith/evaluate-rag-tutorial",
            "https://docs.langchain.com/langsmith/enterprise",
            "https://docs.langchain.com/langsmith/administration-overview",
        ],
    },
    {
        "tool": "Langfuse",
        "category": "Open-source observability plus evaluation platform",
        "recommended_use": "Strong for trace-first RAG quality loops, production scoring, and teams that want self-hosting with documented RBAC and retention controls.",
        "evaluation_posture": "Observability-first evaluation",
        "deployment_posture": "Cloud and self-hosted; enterprise controls vary by edition",
        "rag": 4,
        "ci": 3,
        "dataset": 4,
        "tracing": 5,
        "integration": 4,
        "data_control": 5,
        "governance": 5,
        "evidence": 5,
        "caveat": "Separate OSS/core features from edition-specific enterprise controls before procurement.",
        "source_urls": [
            "https://langfuse.com/guides/cookbook/evaluation_of_rag_with_ragas",
            "https://langfuse.com/docs/administration/rbac",
            "https://langfuse.com/docs/administration/data-retention",
        ],
    },
    {
        "tool": "Phoenix",
        "category": "Open-source observability plus evaluation platform",
        "recommended_use": "Strong for teams that want self-hosted tracing, RAG evaluation, datasets, and privacy controls around evaluation data.",
        "evaluation_posture": "Trace-first RAG evaluation",
        "deployment_posture": "Self-hosted OSS; managed/commercial Arize ecosystem available separately",
        "rag": 4,
        "ci": 3,
        "dataset": 4,
        "tracing": 5,
        "integration": 4,
        "data_control": 5,
        "governance": 4,
        "evidence": 5,
        "caveat": "Managed enterprise controls should be verified separately from self-hosted Phoenix OSS capabilities.",
        "source_urls": [
            "https://arize.com/docs/phoenix/cookbook/evaluation/evaluate-rag",
            "https://arize.com/docs/phoenix/self-hosting/features/authentication",
            "https://arize.com/docs/phoenix/self-hosting/security/privacy",
        ],
    },
    {
        "tool": "Opik",
        "category": "Open-source and managed LLM observability/evaluation",
        "recommended_use": "Good for teams that want tracing, datasets, experiments, online evaluation, and CI hooks in one evaluation family.",
        "evaluation_posture": "Observability plus experiments",
        "deployment_posture": "Cloud and self-hosting documented in project materials",
        "rag": 4,
        "ci": 4,
        "dataset": 5,
        "tracing": 5,
        "integration": 4,
        "data_control": 4,
        "governance": 3,
        "evidence": 4,
        "caveat": "Public research here verified broad capabilities from the official repository; deeper enterprise controls need vendor-doc confirmation.",
        "source_urls": ["https://github.com/comet-ml/opik"],
    },
    {
        "tool": "MLflow GenAI",
        "category": "Open-source AI lifecycle and evaluation platform",
        "recommended_use": "Strong when the organization already standardizes on MLflow and wants GenAI evals, human feedback, tracing, and self-hosted lifecycle governance.",
        "evaluation_posture": "Lifecycle-centered evaluation",
        "deployment_posture": "Fully open-source and self-hostable; managed options depend on platform choice",
        "rag": 3,
        "ci": 4,
        "dataset": 5,
        "tracing": 4,
        "integration": 4,
        "data_control": 5,
        "governance": 4,
        "evidence": 5,
        "caveat": "RAG-specific ergonomics may be less focused than dedicated RAG eval libraries or observability platforms.",
        "source_urls": [
            "https://mlflow.org/docs/latest/genai/eval-monitor/",
            "https://mlflow.org/docs/latest/self-hosting/",
            "https://mlflow.org/docs/latest/self-hosting/security/basic-http-auth/",
        ],
    },
    {
        "tool": "DeepEval / Confident AI",
        "category": "Open-source eval framework plus commercial platform",
        "recommended_use": "Strong for CI/regression tests and unit-test-style LLM app evaluation, with team reporting and governance on the Confident AI platform side.",
        "evaluation_posture": "CI/regression-first evaluation",
        "deployment_posture": "Local OSS runner; cloud/self-hosted enterprise platform",
        "rag": 4,
        "ci": 5,
        "dataset": 4,
        "tracing": 2,
        "integration": 4,
        "data_control": 4,
        "governance": 4,
        "evidence": 4,
        "caveat": "Do not treat enterprise SSO/RBAC/audit/retention as features of the local OSS runner; they are platform-side controls.",
        "source_urls": [
            "https://deepeval.com/docs/getting-started",
            "https://deepeval.com/enterprise",
        ],
    },
    {
        "tool": "Ragas",
        "category": "RAG-focused evaluation framework",
        "recommended_use": "Best when the team needs RAG-specific metrics, test set generation, and framework integrations, and can pair the library with separate experiment tracking.",
        "evaluation_posture": "RAG-metrics-first library",
        "deployment_posture": "Library/local workflow; governance handled by surrounding stack",
        "rag": 5,
        "ci": 3,
        "dataset": 5,
        "tracing": 2,
        "integration": 4,
        "data_control": 3,
        "governance": 2,
        "evidence": 5,
        "caveat": "Excellent evaluator library, not a complete production governance or observability platform by itself.",
        "source_urls": ["https://docs.ragas.io/en/stable/"],
    },
    {
        "tool": "Giskard OSS / Giskard Hub",
        "category": "RAG test suites and LLM security validation",
        "recommended_use": "Good for test-suite-driven RAG validation, out-of-scope handling, multi-turn checks, and organizations that also need AI risk/security workflows.",
        "evaluation_posture": "Test-suite and security oriented",
        "deployment_posture": "OSS library with commercial/on-prem Hub controls",
        "rag": 4,
        "ci": 4,
        "dataset": 4,
        "tracing": 2,
        "integration": 3,
        "data_control": 4,
        "governance": 4,
        "evidence": 4,
        "caveat": "Enterprise controls are Hub-side; verify the OSS workflow against the team's desired CI and observability shape.",
        "source_urls": [
            "https://docs.giskard.ai/oss/checks/use-cases/rag-evaluation",
            "https://www.giskard.ai/",
        ],
    },
    {
        "tool": "W&B Weave",
        "category": "Managed observability and evaluation platform",
        "recommended_use": "Strong for teams already using W&B and wanting traced RAG steps, datasets, scorers, and evaluation UI in the same ecosystem.",
        "evaluation_posture": "Managed tracing plus evaluation",
        "deployment_posture": "Managed SaaS; self-managed enterprise deployment is available but operationally heavy",
        "rag": 4,
        "ci": 3,
        "dataset": 5,
        "tracing": 5,
        "integration": 4,
        "data_control": 4,
        "governance": 4,
        "evidence": 5,
        "caveat": "Self-managed Weave requires an enterprise W&B Platform deployment, Kubernetes, ClickHouse, S3-compatible storage, and a Weave-enabled license.",
        "source_urls": [
            "https://docs.wandb.ai/weave/tutorial-rag",
            "https://docs.wandb.ai/weave/guides/platform/weave-self-managed",
        ],
    },
    {
        "tool": "Promptfoo",
        "category": "CLI/config evaluation runner",
        "recommended_use": "Strong for lightweight local/private evals, prompt/provider comparisons, and CI checks where the team owns the evaluation configuration.",
        "evaluation_posture": "CLI/config-first automation",
        "deployment_posture": "Local/private runner; governance handled by surrounding CI/runtime controls",
        "rag": 3,
        "ci": 5,
        "dataset": 3,
        "tracing": 1,
        "integration": 4,
        "data_control": 5,
        "governance": 2,
        "evidence": 4,
        "caveat": "The OSS runner is not a sandbox for untrusted eval configs or code-executing fields; isolate execution for adversarial packs.",
        "source_urls": [
            "https://www.promptfoo.dev/docs/guides/evaluate-rag/",
            "https://github.com/promptfoo/promptfoo/blob/main/SECURITY.md",
        ],
    },
    {
        "tool": "TruLens",
        "category": "RAG evaluation concepts and feedback functions",
        "recommended_use": "Useful when the team wants explicit RAG Triad-style scoring around context relevance, groundedness, and answer relevance.",
        "evaluation_posture": "RAG-concepts-first evaluation",
        "deployment_posture": "Library/framework workflow; governance handled by surrounding stack",
        "rag": 5,
        "ci": 2,
        "dataset": 3,
        "tracing": 3,
        "integration": 3,
        "data_control": 3,
        "governance": 2,
        "evidence": 4,
        "caveat": "Treat as a strong evaluator/concept layer unless the broader platform requirements are already solved elsewhere.",
        "source_urls": ["https://www.trulens.org/getting_started/core_concepts/rag_triad/"],
    },
    {
        "tool": "Evidently OSS / Platform",
        "category": "Quality monitoring and report-driven evaluation",
        "recommended_use": "Good for notebook/report-driven RAG quality checks and teams that want dataframe outputs, visual reports, and ML quality monitoring patterns.",
        "evaluation_posture": "Report and monitoring oriented",
        "deployment_posture": "OSS library; self-hosted platform; commercial/cloud status needs confirmation",
        "rag": 4,
        "ci": 2,
        "dataset": 3,
        "tracing": 2,
        "integration": 3,
        "data_control": 3,
        "governance": 2,
        "evidence": 3,
        "caveat": "Managed SaaS availability has conflicting official documentation; require vendor confirmation before scoring it as available.",
        "source_urls": [
            "https://docs.evidentlyai.com/examples/LLM_rag_evals",
            "https://docs.evidentlyai.com/faq/oss_vs_cloud",
            "https://docs.evidentlyai.com/faq/cloud_v2",
            "https://docs.evidentlyai.com/docs/setup/self-hosting",
        ],
    },
    {
        "tool": "LlamaIndex evaluators",
        "category": "Framework-native evaluation support",
        "recommended_use": "Use as a primary option only when the production RAG app is built on LlamaIndex; otherwise treat as framework-native support or a metric source.",
        "evaluation_posture": "Framework-native",
        "deployment_posture": "Part of application framework; platform controls come from surrounding stack",
        "rag": 4,
        "ci": 2,
        "dataset": 4,
        "tracing": 2,
        "integration": 5,
        "data_control": 3,
        "governance": 1,
        "evidence": 4,
        "caveat": "Not a standalone evaluation platform for teams outside the LlamaIndex ecosystem.",
        "source_urls": ["https://developers.llamaindex.ai/python/framework/module_guides/evaluating/"],
    },
    {
        "tool": "Haystack evaluation",
        "category": "Framework-native evaluation support",
        "recommended_use": "Use as a primary option only for Haystack-based RAG pipelines; otherwise treat as framework-native evaluation support.",
        "evaluation_posture": "Framework-native",
        "deployment_posture": "Part of application framework; platform controls come from surrounding stack",
        "rag": 4,
        "ci": 2,
        "dataset": 3,
        "tracing": 2,
        "integration": 5,
        "data_control": 3,
        "governance": 1,
        "evidence": 4,
        "caveat": "Strongest inside Haystack pipelines; not a standalone observability or governance platform.",
        "source_urls": ["https://docs.haystack.deepset.ai/docs/model-based-evaluation"],
    },
    {
        "tool": "OpenAI Evals / Evals API",
        "category": "OpenAI-specific legacy/caveat",
        "recommended_use": "Do not choose as a long-term primary RAG evaluation stack; keep only as a legacy/OpenAI-specific caveat where already in use.",
        "evaluation_posture": "Deprecated/legacy",
        "deployment_posture": "OpenAI platform-specific; long-term availability risk",
        "rag": 2,
        "ci": 2,
        "dataset": 3,
        "tracing": 1,
        "integration": 3,
        "data_control": 1,
        "governance": 1,
        "evidence": 5,
        "caveat": "Official docs state Evals platform read-only for existing users on 2026-10-31 and shutdown on 2026-11-30.",
        "source_urls": ["https://developers.openai.com/api/docs/guides/evals"],
    },
]


WEIGHTS = [
    ("RAG-specific metrics and evaluators", 20, "Evaluation", "Does the tool directly evaluate retrieval quality, groundedness, answer quality, and related RAG failure modes?"),
    ("Repeatable regression and CI workflow", 15, "Evaluation", "How directly can teams run repeatable eval suites in CI or automated regression gates?"),
    ("Dataset/testset and experiment workflow", 15, "Evaluation", "How well does the tool support datasets, generated test sets, experiments, examples, or runs?"),
    ("Observability/tracing and production feedback loop", 15, "Evaluation", "Can scores be attached to traces/spans or production feedback loops?"),
    ("Integration fit with common RAG stacks", 10, "Evaluation", "How cleanly does it integrate with common RAG frameworks and app stacks?"),
    ("Deployment/data-control posture", 10, "Operational", "Is the tool local, self-hostable, hybrid, or otherwise usable under data-control constraints?"),
    ("Access control, retention, audit, and enterprise governance", 10, "Operational", "Are RBAC, retention, audit, SSO, or equivalent controls publicly documented?"),
    ("Public evidence maturity", 5, "Both", "How strong and specific is the public evidence used for this row?"),
]


CAVEATS = [
    ("Evidently managed SaaS", "Current official docs conflict: some pages say Cloud is no longer available as SaaS, while Cloud v2 docs describe SaaS onboarding. Treat as vendor-confirmation required.", "Blocking for teams that require managed SaaS."),
    ("OpenAI Evals", "Official OpenAI docs mark the Evals platform as deprecated, with read-only status on 2026-10-31 and shutdown on 2026-11-30.", "Do not select as a long-term primary stack."),
    ("Framework-native rows", "LlamaIndex and Haystack evaluators are valuable when the app already uses those frameworks, but they are not standalone observability/governance platforms.", "Use as framework support, not a default cross-stack choice."),
    ("Commercial governance claims", "Some enterprise controls are documented on vendor pages rather than deep implementation docs.", "Use evidence confidence and procurement/security review before adoption."),
    ("Promptfoo execution model", "Promptfoo OSS is a local eval runner, not a sandbox for untrusted configs or code-executing eval fields.", "Run untrusted eval packs in isolated CI/runtime environments."),
    ("Scores are fit scores, not benchmark results", "The matrix scores publicly documented capabilities and production fit; it does not benchmark model quality or runtime performance.", "Run a project-specific proof of concept before final tool selection."),
]


def read_jsonl(path: Path) -> list[dict]:
    records = []
    if not path.exists():
        return records
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            records.append(json.loads(line))
    return records


def normalize_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "", name)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T{cleaned}"
    return cleaned[:31]


def source_map(sources: list[dict]) -> tuple[dict[str, str], list[dict]]:
    by_url = {}
    ordered = []
    for idx, record in enumerate(sources, start=1):
        url = record.get("url", "")
        if not url or url in by_url:
            continue
        source_id = f"SRC{idx:03d}"
        by_url[url] = source_id
        ordered.append(
            {
                "source_id": source_id,
                "title": record.get("title", "").strip(),
                "url": url,
                "note": record.get("note", "").strip(),
                "recorded_at": record.get("recorded_at", ""),
            }
        )
    return by_url, ordered


def score_rows() -> list[dict]:
    rows = []
    for item in TOOLS:
        eval_score = round(
            ((item["rag"] * 20) + (item["ci"] * 15) + (item["dataset"] * 15) + (item["tracing"] * 15) + (item["integration"] * 10) + (item["evidence"] * 5))
            / (5 * 80)
            * 100
        )
        op_score = round(((item["data_control"] * 10) + (item["governance"] * 10) + (item["evidence"] * 5)) / (5 * 25) * 100)
        overall = round((eval_score * 0.7) + (op_score * 0.3))
        if "Deprecated/legacy" in item["evaluation_posture"]:
            tier = "Avoid/legacy"
        elif overall >= 85:
            tier = "Adopt/shortlist"
        elif overall >= 70:
            tier = "Strong candidate"
        elif overall >= 55:
            tier = "Situational"
        elif overall >= 40:
            tier = "Caveat/secondary"
        else:
            tier = "Avoid/legacy"
        rows.append({**item, "eval_score": eval_score, "op_score": op_score, "overall": overall, "tier": tier})
    return sorted(rows, key=lambda r: (r["overall"], r["eval_score"]), reverse=True)


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_table(ws, name: str, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    table = Table(displayName=normalize_table_name(name), ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def set_widths(ws, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_workbook(rows: list[dict], sources: list[dict], by_url: dict[str, str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["RAG Evaluation Tooling Matrix", "Generated", GENERATED_ON])
    ws["A1"].font = Font(size=16, bold=True)
    ws["B1"].font = Font(bold=True)
    ws.append([])
    ws.append(["Recommendation", "Tool family", "Why it matters", "Key caveat"])
    top_rows = rows[:7]
    for row in top_rows:
        ws.append([row["tier"], row["tool"], row["recommended_use"], row["caveat"]])
    style_header(ws, 3)
    add_table(ws, "SummaryTable", 3, 3 + len(top_rows), 1, 4)
    ws.freeze_panes = "A4"
    set_widths(ws, {1: 20, 2: 28, 3: 78, 4: 68})
    for row in ws.iter_rows(min_row=4, max_row=3 + len(top_rows), min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    matrix = wb.create_sheet("Tool Matrix")
    headers = [
        "Tool family",
        "Category",
        "Recommended use",
        "Evaluation posture",
        "Deployment posture",
        "RAG metrics",
        "CI/regression",
        "Datasets/testsets",
        "Tracing/feedback",
        "Integration fit",
        "Data control",
        "Governance",
        "Evidence maturity",
        "Evaluation Fit Score",
        "Operational Fit Score",
        "Overall Score",
        "Recommendation Tier",
        "Key caveat",
        "Primary source IDs",
    ]
    matrix.append(headers)
    for excel_row, row in enumerate(rows, start=2):
        source_ids = [by_url.get(url, "unlisted") for url in row["source_urls"]]
        matrix.append(
            [
                row["tool"],
                row["category"],
                row["recommended_use"],
                row["evaluation_posture"],
                row["deployment_posture"],
                row["rag"],
                row["ci"],
                row["dataset"],
                row["tracing"],
                row["integration"],
                row["data_control"],
                row["governance"],
                row["evidence"],
                f"=ROUND(((F{excel_row}*20)+(G{excel_row}*15)+(H{excel_row}*15)+(I{excel_row}*15)+(J{excel_row}*10)+(M{excel_row}*5))/(5*80)*100,0)",
                f"=ROUND(((K{excel_row}*10)+(L{excel_row}*10)+(M{excel_row}*5))/(5*25)*100,0)",
                f"=ROUND((N{excel_row}*0.7)+(O{excel_row}*0.3),0)",
                f'=IF(D{excel_row}="Deprecated/legacy","Avoid/legacy",IF(P{excel_row}>=85,"Adopt/shortlist",IF(P{excel_row}>=70,"Strong candidate",IF(P{excel_row}>=55,"Situational",IF(P{excel_row}>=40,"Caveat/secondary","Avoid/legacy")))))',
                row["caveat"],
                ", ".join(source_ids),
            ]
        )
    style_header(matrix)
    add_table(matrix, "ToolMatrix", 1, len(rows) + 1, 1, len(headers))
    matrix.freeze_panes = "A2"
    set_widths(
        matrix,
        {
            1: 26,
            2: 34,
            3: 70,
            4: 24,
            5: 42,
            6: 12,
            7: 14,
            8: 16,
            9: 16,
            10: 14,
            11: 13,
            12: 12,
            13: 16,
            14: 18,
            15: 18,
            16: 14,
            17: 20,
            18: 68,
            19: 24,
        },
    )
    dv_score = DataValidation(type="whole", operator="between", formula1="0", formula2="5", allow_blank=False)
    matrix.add_data_validation(dv_score)
    dv_score.add(f"F2:M{len(rows)+1}")
    matrix.conditional_formatting.add(f"N2:P{len(rows)+1}", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=65, mid_color="FFEB84", end_type="num", end_value=100, end_color="63BE7B"))
    for row in matrix.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    scoring = wb.create_sheet("Scoring")
    scoring.append(["Criterion", "Weight", "Layer", "Notes"])
    for weight in WEIGHTS:
        scoring.append(list(weight))
    scoring.append([])
    scoring.append(["Scoring scale", "Meaning", "Use"])
    scale_rows = [
        (5, "Strong, directly documented fit", "Use when public docs show the capability as a first-class workflow."),
        (4, "Good documented fit", "Use when docs show the capability but with narrower scope or edition caveats."),
        (3, "Moderate fit", "Use when capability exists but depends on surrounding stack or extra work."),
        (2, "Limited fit", "Use when public docs imply partial support or weak operational fit."),
        (1, "Minimal or governance handled elsewhere", "Use for framework/library rows without platform controls."),
        (0, "Not found or not applicable", "Use when public evidence is absent or the feature is clearly out of scope."),
    ]
    for scale in scale_rows:
        scoring.append(list(scale))
    style_header(scoring)
    style_header(scoring, len(WEIGHTS) + 3)
    add_table(scoring, "WeightsTable", 1, len(WEIGHTS) + 1, 1, 4)
    add_table(scoring, "ScaleTable", len(WEIGHTS) + 3, len(WEIGHTS) + 3 + len(scale_rows), 1, 3)
    set_widths(scoring, {1: 44, 2: 12, 3: 16, 4: 86})
    scoring.freeze_panes = "A2"

    ev = wb.create_sheet("Evidence Sources")
    ev.append(["Source ID", "Title", "URL", "Evidence note", "Recorded at"])
    for source in sources:
        ev.append([source["source_id"], source["title"], source["url"], source["note"], source["recorded_at"]])
    style_header(ev)
    add_table(ev, "EvidenceSources", 1, len(sources) + 1, 1, 5)
    ev.freeze_panes = "A2"
    set_widths(ev, {1: 12, 2: 42, 3: 72, 4: 92, 5: 22})
    for row in ev.iter_rows(min_row=2, max_row=len(sources) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    caveats = wb.create_sheet("Exclusions Caveats")
    caveats.append(["Topic", "Caveat", "Decision impact"])
    for caveat in CAVEATS:
        caveats.append(list(caveat))
    style_header(caveats)
    add_table(caveats, "CaveatsTable", 1, len(CAVEATS) + 1, 1, 3)
    caveats.freeze_panes = "A2"
    set_widths(caveats, {1: 28, 2: 90, 3: 52})
    for row in caveats.iter_rows(min_row=2, max_row=len(CAVEATS) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    method = wb.create_sheet("Methodology")
    method_rows = [
        ("Purpose", "Help an engineering team shortlist RAG/LLM application evaluation tooling for a production RAG project."),
        ("Source policy", "Official documentation, GitHub repositories, release/vendor documentation, and public vendor docs were preferred over blog summaries."),
        ("Safety policy", "The package contains only public-source evidence and does not include personal/contact data, credentials, or local environment identifiers."),
        ("Evaluation Fit Score", "Weighted 0-100 score over RAG metrics, CI/regression, datasets/testsets, tracing/feedback, integration fit, and evidence maturity."),
        ("Operational Fit Score", "Weighted 0-100 score over deployment/data-control posture, governance controls, and evidence maturity."),
        ("Recommendation Tier", "Formula-derived tier from the overall score, with deprecated/legacy tools explicitly capped as Avoid/legacy."),
        ("Important limitation", "Scores are decision-support fit scores based on public evidence. They are not benchmark results and should be validated with a project-specific proof of concept."),
        ("Generated on", GENERATED_ON),
    ]
    method.append(["Item", "Description"])
    for item in method_rows:
        method.append(list(item))
    style_header(method)
    add_table(method, "MethodologyTable", 1, len(method_rows) + 1, 1, 2)
    method.freeze_panes = "A2"
    set_widths(method, {1: 28, 2: 112})
    for row in method.iter_rows(min_row=2, max_row=len(method_rows) + 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(WORKBOOK_PATH)


def md_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    header = rows[0]
    out = ["| " + " | ".join(header) + " |", "| " + " | ".join(["---"] * len(header)) + " |"]
    for row in rows[1:]:
        escaped = [str(cell).replace("\n", "<br>").replace("|", "\\|") for cell in row]
        out.append("| " + " | ".join(escaped) + " |")
    return "\n".join(out)


def write_markdown(rows: list[dict], sources: list[dict]) -> None:
    top = rows[:7]
    caveat_rows = [[topic, caveat, impact] for topic, caveat, impact in CAVEATS]
    tool_rows = [
        ["Tool family", "Tier", "Eval fit", "Operational fit", "Best use", "Main caveat"]
    ] + [
        [r["tool"], r["tier"], str(r["eval_score"]), str(r["op_score"]), r["recommended_use"], r["caveat"]]
        for r in rows
    ]

    readme = f"""# RAG Evaluation Tooling Matrix

This example package compares current RAG and LLM application evaluation tooling for a production RAG project. It is designed as a compact public-safe decision aid: every tool row is backed by public evidence links, and the workbook separates evaluation capability from deployment and governance fit.

## Included files

- `rag-eval-tooling-matrix.xlsx` - Excel workbook with Summary, Tool Matrix, Scoring, Evidence Sources, Exclusions Caveats, and Methodology sheets.
- `final-report.md` - human-readable analysis and recommendation notes.
- `sources.md` - public evidence index used by the workbook.
- `validation-report.md` - workbook structure and compatibility checks from generation time.

## How to use the workbook

1. Start in the Summary sheet for the shortlist.
2. Use Tool Matrix to compare evaluation fit, operational fit, and caveats.
3. Adjust 0-5 criterion scores if your project has different constraints.
4. Review Evidence Sources and Exclusions Caveats before treating a score as procurement-ready.

The matrix is intentionally not a benchmark. It scores documented fit, public evidence maturity, and production adoption risk.

## Short take

- For managed or enterprise-heavy teams, start with LangSmith, Langfuse, Phoenix, W&B Weave, Opik, or MLflow GenAI depending on your existing stack.
- For CI/regression-first workflows, compare DeepEval / Confident AI and Promptfoo.
- For RAG metric depth, consider Ragas and TruLens, usually paired with a separate tracking or observability layer.
- For framework-native projects, LlamaIndex and Haystack evaluators are useful when the app already uses those frameworks.
- Treat OpenAI Evals / Evals API as legacy because official OpenAI docs now describe a deprecation and shutdown timeline.

Generated on {GENERATED_ON} from public sources.
"""

    report = f"""# RAG and LLM Application Evaluation Tooling Matrix

## Executive summary

The strongest production choices depend less on a single universal "best" evaluator and more on the team's operating model:

- **Managed experiment and governance workflow:** LangSmith is the clearest fit for LangChain-heavy teams that want datasets, experiment runs, traces, and enterprise deployment options in one platform.
- **Open-source trace-first quality loop:** Langfuse and Phoenix are strong when self-hosting, data control, and trace-level evaluation matter.
- **Lifecycle platform fit:** MLflow GenAI is attractive for teams already standardizing on MLflow and wanting evaluation, feedback, tracing, and self-hosted governance in one lifecycle layer.
- **CI/regression gate:** DeepEval / Confident AI and Promptfoo are strong options for repeatable test suites, with Promptfoo being lighter and more config-oriented.
- **RAG-specific metric layer:** Ragas and TruLens are useful evaluator layers, but they usually need to be paired with experiment tracking, trace collection, or governance tooling.

## Shortlist table

{md_table(tool_rows)}

## Scoring model

The workbook uses two visible scoring layers:

- **Evaluation Fit Score:** RAG-specific metrics, CI/regression support, datasets/testsets, observability/tracing, integration fit, and public evidence maturity.
- **Operational Fit Score:** deployment/data-control posture, access control/retention/audit/governance, and public evidence maturity.

The default weighted criteria are:

{md_table([["Criterion", "Weight", "Layer", "Notes"]] + [[c, str(w), layer, notes] for c, w, layer, notes in WEIGHTS])}

## Important caveats

{md_table([["Topic", "Caveat", "Decision impact"]] + caveat_rows)}

## Recommended adoption path

1. Pick one trace/observability-backed option and one CI/regression option for a two-tool proof of concept.
2. Run both against the same representative RAG failures: irrelevant retrieval, unsupported answer, stale source, out-of-scope user query, and multi-turn context drift.
3. Require evidence capture: prompt, retrieved context, answer, expected behavior, scores, source links, and reviewer notes.
4. Before procurement, validate retention, RBAC, SSO, audit logging, data residency, and self-hosting/SaaS status directly with current vendor documentation or security review.

## Evidence policy

This package uses public official documentation, GitHub repositories, release/vendor documentation, and vendor docs. It does not include confidential identifiers, credentials, or contact data.
"""

    source_lines = ["# Sources", "", f"Generated on {GENERATED_ON}.", ""]
    for source in sources:
        source_lines.extend(
            [
                f"## {source['source_id']} - {source['title']}",
                "",
                f"- URL: {source['url']}",
                f"- Evidence note: {source['note']}",
                "",
            ]
        )

    (OUTPUT_DIR / "README.md").write_text(readme, encoding="utf-8")
    (OUTPUT_DIR / "final-report.md").write_text(report, encoding="utf-8")
    (OUTPUT_DIR / "sources.md").write_text("\n".join(source_lines), encoding="utf-8")


def ranges_overlap(a: str, b: str) -> bool:
    a_min_col, a_min_row, a_max_col, a_max_row = range_boundaries(a)
    b_min_col, b_min_row, b_max_col, b_max_row = range_boundaries(b)
    return not (a_max_col < b_min_col or b_max_col < a_min_col or a_max_row < b_min_row or b_max_row < a_min_row)


def validate_workbook() -> tuple[bool, list[str]]:
    messages = []
    ok = True
    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    expected_sheets = {"Summary", "Tool Matrix", "Scoring", "Evidence Sources", "Exclusions Caveats", "Methodology"}
    missing = expected_sheets - set(wb.sheetnames)
    if missing:
        ok = False
        messages.append(f"Missing sheets: {', '.join(sorted(missing))}")
    else:
        messages.append("All expected worksheets are present.")

    for ws in wb.worksheets:
        refs = []
        for table_name in ws.tables:
            table = ws.tables[table_name]
            refs.append((table_name, table.ref))
        for idx, (name_a, ref_a) in enumerate(refs):
            for name_b, ref_b in refs[idx + 1 :]:
                if ranges_overlap(ref_a, ref_b):
                    ok = False
                    messages.append(f"Table ranges overlap on {ws.title}: {name_a} {ref_a} and {name_b} {ref_b}.")
        sheet_filter = ws.auto_filter.ref
        if sheet_filter:
            for name, ref in refs:
                if ranges_overlap(sheet_filter, ref):
                    ok = False
                    messages.append(f"Worksheet autoFilter overlaps table range on {ws.title}: {sheet_filter} and {name} {ref}.")
        if refs:
            messages.append(f"{ws.title}: {len(refs)} Excel table(s), no worksheet-level autoFilter overlap detected.")
        else:
            messages.append(f"{ws.title}: no Excel tables.")

    formulas = 0
    for row in wb["Tool Matrix"].iter_rows(min_row=2, max_col=17):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas += 1
    if formulas < 3 * len(TOOLS):
        ok = False
        messages.append(f"Expected matrix formulas were not fully present; found {formulas}.")
    else:
        messages.append(f"Tool Matrix contains {formulas} formula cells.")

    report = ["# Workbook validation report", "", f"Workbook: `{WORKBOOK_PATH.name}`", "", f"Status: {'passed' if ok else 'failed'}", ""]
    report.extend(f"- {message}" for message in messages)
    VALIDATION_PATH.write_text("\n".join(report) + "\n", encoding="utf-8")
    return ok, messages


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stale_public_summary = OUTPUT_DIR / "generation-summary.json"
    if stale_public_summary.exists():
        stale_public_summary.unlink()
    raw_sources = read_jsonl(SOURCES_PATH)
    by_url, sources = source_map(raw_sources)
    rows = score_rows()
    create_workbook(rows, sources, by_url)
    write_markdown(rows, sources)
    ok, messages = validate_workbook()

    summary = {
        "package_dir": str(OUTPUT_DIR),
        "workbook": str(WORKBOOK_PATH),
        "validation_report": str(VALIDATION_PATH),
        "validation_passed": ok,
        "tool_rows": len(rows),
        "source_rows": len(sources),
        "validation_messages": messages,
    }
    GENERATION_SUMMARY_PATH.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    if not ok:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
